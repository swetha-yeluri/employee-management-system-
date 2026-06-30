
import csv
import io

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.controllers import analytics_controller
from app.models.attendance_record_model import AttendanceRecord
from app.models.audit_log_model import AuditLog
from app.models.employee_model import Employee
from app.models.export_history_model import ExportHistory
from app.models.leave_request_model import LeaveRequest
from app.models.notification_model import Notification
from app.models.user_model import User

DATA_TYPES = ["employees", "attendance", "leaves", "audit-logs", "notifications", "analytics"]
FORMATS = ["csv", "excel", "pdf"]

_MEDIA = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
_EXT = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}


def _fmt(dt):
    return dt.isoformat(sep=" ", timespec="seconds") if dt else ""


def _user_emails(db: Session, company_id: int) -> dict:
    return {u.id: u.email for u in db.query(User).filter(User.company_id == company_id).all()}


def gather(db: Session, company_id: int, data_type: str):
    """Returns (title, headers, rows) for the given company-scoped data type."""
    if data_type == "employees":
        rows = []
        for e in db.query(Employee).filter(Employee.company_id == company_id).order_by(Employee.id).all():
            rows.append([e.id, e.name, e.email, e.position, e.status,
                         e.department.name if e.department else "Unassigned"])
        return "Employees", ["ID", "Name", "Email", "Position", "Status", "Department"], rows

    if data_type == "attendance":
        emails = _user_emails(db, company_id)
        rows = []
        for r in db.query(AttendanceRecord).filter(AttendanceRecord.company_id == company_id).order_by(AttendanceRecord.check_in.desc()).all():
            rows.append([emails.get(r.user_id, r.user_id), r.work_date,
                         _fmt(r.check_in), _fmt(r.check_out), r.work_hours if r.work_hours is not None else ""])
        return "Attendance", ["User", "Date", "Check In", "Check Out", "Hours"], rows

    if data_type == "leaves":
        rows = []
        for l in db.query(LeaveRequest).filter(LeaveRequest.company_id == company_id).order_by(LeaveRequest.created_at.desc()).all():
            rows.append([l.user_email, l.leave_type, l.start_date, l.end_date, l.reason, l.status])
        return "Leave Requests", ["User", "Type", "Start", "End", "Reason", "Status"], rows

    if data_type == "audit-logs":
        rows = []
        for a in db.query(AuditLog).filter(AuditLog.company_id == company_id).order_by(AuditLog.timestamp.desc()).all():
            rows.append([a.user_name, a.action, a.target, _fmt(a.timestamp)])
        return "Audit Logs", ["User", "Action", "Target", "Timestamp"], rows

    if data_type == "notifications":
        emails = _user_emails(db, company_id)
        rows = []
        for n in db.query(Notification).filter(Notification.company_id == company_id).order_by(Notification.created_at.desc()).all():
            rows.append([emails.get(n.user_id, n.user_id), n.message,
                         "Read" if n.is_read else "Unread", _fmt(n.created_at)])
        return "Notifications", ["User", "Message", "Status", "Created"], rows

    if data_type == "analytics":
        a = analytics_controller.get_analytics(db, company_id)
        rows = [
            ["Total Employees", a["total_employees"]],
            ["Active Employees", a["active_employees"]],
            ["Total Departments", a["total_departments"]],
            ["Pending Requests", a["pending_requests"]],
        ]
        for item in a["by_department"]:
            rows.append([f"Department: {item['label']}", item["count"]])
        for item in a["by_status"]:
            rows.append([f"Status: {item['label']}", item["count"]])
        return "Analytics", ["Metric", "Value"], rows

    raise HTTPException(400, "Unknown data type")


# ---------- renderers ----------
def _to_csv(headers, rows) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for row in rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8")


def _to_excel(headers, rows, title) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "Excel export needs openpyxl. In backend venv run: pip install openpyxl fpdf2")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append([("" if c is None else c) for c in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_pdf(headers, rows, title) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(500, "PDF export needs fpdf2. In backend venv run: pip install openpyxl fpdf2")
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
    n = max(len(headers), 1)
    usable = pdf.w - 2 * pdf.l_margin
    col_w = usable / n
    pdf.set_font("Helvetica", "B", 8)
    for h in headers:
        pdf.cell(col_w, 7, str(h)[:28], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for row in rows:
        for c in row:
            pdf.cell(col_w, 6, str("" if c is None else c)[:28], border=1)
        pdf.ln()
    out = pdf.output()
    return bytes(out)


def render(data_type, file_format, title, headers, rows) -> bytes:
    if file_format == "csv":
        return _to_csv(headers, rows)
    if file_format == "excel":
        return _to_excel(headers, rows, title)
    if file_format == "pdf":
        return _to_pdf(headers, rows, title)
    raise HTTPException(400, "Unknown format")


def export(db: Session, admin, data_type: str, file_format: str):
    if data_type not in DATA_TYPES:
        raise HTTPException(400, "Unknown data type")
    if file_format not in FORMATS:
        raise HTTPException(400, "Unknown format")

    title, headers, rows = gather(db, admin.company_id, data_type)
    content = render(data_type, file_format, title, headers, rows)

    db.add(ExportHistory(company_id=admin.company_id, exported_by=admin.email,
                         data_type=data_type, file_format=file_format))
    db.commit()

    filename = f"{data_type}.{_EXT[file_format]}"
    return content, _MEDIA[file_format], filename


def list_history(db: Session, admin):
    return (
        db.query(ExportHistory)
        .filter(ExportHistory.company_id == admin.company_id)
        .order_by(ExportHistory.created_at.desc())
        .limit(50)
        .all()
    )
